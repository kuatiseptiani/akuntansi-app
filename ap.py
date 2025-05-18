import io
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import streamlit as st
import pandas as pd
import locale
locale.setlocale(locale.LC_ALL, '')

st.set_page_config(page_title="Aplikasi CV Bee Alaska", layout="wide")

# Inisialisasi session state
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
if "users" not in st.session_state:
    st.session_state.users = {"admin": "1"}
if "username" not in st.session_state:
    st.session_state.username = ""
if "show_notification" not in st.session_state:
    st.session_state.show_notification = False
if "notification_message" not in st.session_state:
    st.session_state.notification_message = ""
if "notification_type" not in st.session_state:
    st.session_state.notification_type = "success"

# Fungsi untuk menampilkan notifikasi
def show_notification():
    if st.session_state.show_notification:
        notification_container = st.container()
        
        # Buat tombol close dan pesan dalam satu baris
        col1, col2 = notification_container.columns([0.95, 0.05])
        
        # Tampilkan pesan sesuai tipe
        if st.session_state.notification_type == "success":
            col1.success(st.session_state.notification_message)
        elif st.session_state.notification_type == "info":
            col1.info(st.session_state.notification_message)
        elif st.session_state.notification_type == "warning":
            col1.warning(st.session_state.notification_message)
        elif st.session_state.notification_type == "error":
            col1.error(st.session_state.notification_message)
        
        # Tombol untuk menutup notifikasi
        if col2.button("✖"):
            st.session_state.show_notification = False
            st.rerun()

# Panggil fungsi notifikasi di awal aplikasi
show_notification()

# Autentikasi login
def login():
    st.sidebar.subheader("🔐 Halaman Login")
    username = st.sidebar.text_input("Username")
    password = st.sidebar.text_input("Password", type="password")
    
    if st.sidebar.button("Login"):
        if username in st.session_state.users and st.session_state.users[username] == password:
            st.session_state.logged_in = True
            st.session_state.username = username
            
            # Set notifikasi login berhasil
            st.session_state.show_notification = True
            st.session_state.notification_message = f"Login berhasil! Selamat datang, {username}!"
            st.session_state.notification_type = "success"
            
            st.rerun()
        else:
            st.sidebar.error("Username atau password salah")

# Jika belum login, panggil form login dan hentikan
if not st.session_state.logged_in:
    login()
    st.stop()

# sidebar
st.sidebar.title("Siklus Akuntansi")
menu = st.sidebar.radio("", ["Neraca Saldo", "Jurnal Umum", "Buku Besar", 
                             "Neraca Saldo Setelah Disesuaikan", "Laporan Laba Rugi", 
                             "Laporan Perubahan Ekuitas", "Laporan Posisi Keuangan"])

# Tombol logout
if st.sidebar.button("Logout"):
    username = st.session_state.username  # Simpan username untuk pesan
    st.session_state.logged_in = False
    st.session_state.username = ""
    
    # Set notifikasi logout berhasil
    st.session_state.show_notification = True
    st.session_state.notification_message = f"Logout berhasil! Sampai jumpa kembali, {username}!"
    st.session_state.notification_type = "info"
    
    st.rerun()

#Neraca Saldo
if menu == "Neraca Saldo":
    st.title("🐝 CV Bee Alaska")
    st.subheader("Neraca Saldo")
    df_akun = ["Kas","Piutang Usaha", "Perlengkapan", "Persediaan Pakan", "Persediaan Madu", "Peralatan usaha", "Kendaraan", "Akumulasi Penyusutan Kendaraan", "Utang Usaha", "Utang Gaji", "Utang Pajak", "Modal, Alwi", "Prive, Alwi", "Ikhtisar Laba Rugi", "Penjualan", "Harga Pokok Penjualan", "Beban Gaji Karyawan", "Beban Pakan Lebah", "Beban Perlengkapan", "Beban Perawatan Sarang", "Beban Penyusutan Kendaraan", "Beban lain-lain", "Beban Pajak"]
    
    if "saldo_awal" not in st.session_state:
        st.session_state["saldo_awal"] = []
    
    with st.expander("Neraca Saldo"):
        akun = st.selectbox("Nama Akun", df_akun)
        #keterangan = st.selectbox("Keterangan", ["Kas", "Prive", "Utang", "Modal", "Pendapatan", "Beban"])
        saldo_debit = st.number_input("Saldo Debit", min_value=0.0)
        saldo_kredit = st.number_input("Saldo Kredit", min_value=0.0)
        if st.button("Tambah Saldo Awal"):
            st.session_state["saldo_awal"].append({
                "Nama Akun": akun,
                #"Keterangan": keterangan,
                "Debit": saldo_debit,
                "Kredit": saldo_kredit
            })
            st.success("Saldo awal berhasil ditambahkan!")
            
    # Tampilkan Neraca Saldo Awal
    #st.subheader("📄 Neraca Saldo")
    df = pd.DataFrame(st.session_state["saldo_awal"])
    
    if not df.empty and "Debit" in df.columns and "Kredit" in df.columns:
        saldo_debit = df["Debit"].sum()
        saldo_kredit = df["Kredit"].sum()
        
        total_row = pd.DataFrame([{
            "Nama Akun": ""   "Total",
            #"Keterangan": keterangan,
            "Debit": saldo_debit,
            "Kredit": saldo_kredit
        }])
        df_total = pd.concat([df, total_row], ignore_index=True)
        
        def format_rupiah(x):
            if x == 0 or x == "":
                return ""
            return f"Rp {int(x):,}".replace(",", ".")
        
        df_total["Debit"] = df_total["Debit"].apply(format_rupiah)
        df_total["Kredit"] = df_total["Kredit"].apply(format_rupiah)
        
        st.dataframe(df_total, use_container_width=True)
    else:
        st.info("Belum ada data nerasa saldo. Silakan tambahkan dulu.")
    
    if saldo_debit != saldo_kredit:
            st.warning("⚠️ Total Debit tidak sama dengan Total Kredit!")
            
    # Simpan ke Excel

#Jurnal Umum
if menu == "Jurnal Umum":
    st.title("🐝 CV Bee Alaska")
    st.subheader("Jurnal Umum")
    df_akun = ["Kas","Piutang Usaha", "Perlengkapan", "Persediaan Pakan", "Persediaan Madu", "Peralatan usaha", "Kendaraan", "Akumulasi Penyusutan Kendaraan", "Utang Usaha", "Utang Gaji", "Utang Pajak", "Modal, Alwi", "Prive, Alwi", "Ikhtisar laba rugi", "Penjualan", "Harga Pokok Penjualan", "Beban Gaji Karyawan", "Beban Pakan Lebah", "Beban Perlengkapan", "Beban Perawatan Sarang", "Beban Penyusutan Kendaraan", "Beban lain-lain", "Beban Pajak"]
    
    # Inisialisasi data
    if "jurnal" not in st.session_state:
        st.session_state["jurnal"] = []
        
    # Form input transaksi
    with st.expander("Tambah Transaksi"):
        tanggal = st.date_input("Tanggal")
        nama_akun = st.selectbox("Nama Akun",df_akun)
        keterangan = st.text_input("Keterangan")
        saldo_debit = st.number_input("Saldo Debit", min_value=0.0)
        saldo_kredit = st.number_input("Saldo Kredit", min_value=0.0)
        if st.button("Tambah Transaksi"):
            st.session_state["jurnal"].append({
                "Tanggal": str(tanggal),
                "Nama Akun": nama_akun,
                "Keterangan": keterangan,
                "Debit": saldo_debit,
                "Kredit": saldo_kredit
            })
            st.success("Transaksi berhasil ditambahkan!")
            
    # Tampilkan jurnal umum
    #st.subheader("📄 Jurnal Umum")
    df = pd.DataFrame(st.session_state["jurnal"])
    
    if not df.empty and "Debit" in df.columns and "Kredit" in df.columns:
        total_debit = df["Debit"].sum()
        total_kredit = df["Kredit"].sum()
        
        total_row = pd.DataFrame([{
            "Tanggal": "",
            "Nama Akun": "",
            "Keterangan": "" "Total",
            "Debit": total_debit,
            "Kredit": total_kredit
        }])
        df_total = pd.concat([df, total_row], ignore_index=True)
        
        def format_rupiah(x):
            if x == 0 or x == "":
                return ""
            return f"Rp {int(x):,}".replace(",", ".")
        
        df_total["Debit"] = df_total["Debit"].apply(format_rupiah)
        df_total["Kredit"] = df_total["Kredit"].apply(format_rupiah)
        
        st.dataframe(df_total, use_container_width=True)
    else:
        st.info("Belum ada data transaksi. Silakan tambahkan dulu.")
        
    if saldo_debit != saldo_kredit:
            st.warning("⚠️ Total Debit tidak sama dengan Total Kredit!")
    
    # Simpan ke Excel
        
#Buku Besar
if menu == "Buku Besar":
    st.title("🐝 CV Bee Alaska")
    st.subheader("Buku Besar")

    # Kelompok akun
    akun_debit_normal = ["Kas","Piutang Usaha", "Perlengkapan", "Persediaan Pakan", "Persediaan Madu", "Peralatan usaha", "Kendaraan", "Prive, Alwi", "Harga Pokok Penjualan", "Beban Gaji Karyawan", "Beban Pakan Lebah", "Beban Perlengkapan", "Beban Perawatan Sarang", "Beban Penyusutan Kendaraan", "Beban lain-lain", "Beban Pajak"]
    akun_kredit_normal = ["Akumulasi Penyusutan Kendaraan", "Utang Usaha", "Utang Gaji", "Utang Pajak", "Modal, Alwi", "Penjualan"]

    # Fungsi saldo awal
    def saldo_awal_akun(akun, debit, kredit):
        if akun in akun_debit_normal:
            return + debit - kredit
        elif akun in akun_kredit_normal:
            return + kredit - debit
        else:
            return + debit - kredit

    # Fungsi saldo berjalan
    def hitung_saldo_berjalan(akun, debit, kredit, saldo_sebelumnya):
        if akun in akun_debit_normal:
            return saldo_sebelumnya + debit - kredit
        elif akun in akun_kredit_normal:
            return saldo_sebelumnya + kredit - debit
        else:
            return saldo_sebelumnya + debit - kredit

    # Ambil data dari session_state
    saldo_awal = st.session_state.get("saldo_awal", [])
    jurnal = st.session_state.get("jurnal", [])

    # Buat dict saldo awal
    saldo_awal_dict = {}
    for item in saldo_awal:
        nama = item["Nama Akun"]
        saldo_awal_dict[nama] = saldo_awal_akun(nama, item["Debit"], item["Kredit"])

    # Persiapkan jurnal
    df_jurnal = pd.DataFrame(jurnal)
    if not df_jurnal.empty:
        df_jurnal["Tanggal"] = pd.to_datetime(df_jurnal["Tanggal"], errors='coerce')
        df_jurnal["Debit"] = pd.to_numeric(df_jurnal["Debit"], errors='coerce').fillna(0)
        df_jurnal["Kredit"] = pd.to_numeric(df_jurnal["Kredit"], errors='coerce').fillna(0)
        df_jurnal = df_jurnal.sort_values("Tanggal").reset_index(drop=True)
    else:
        df_jurnal = pd.DataFrame(columns=["Tanggal", "Nama Akun", "Keterangan", "Debit", "Kredit"])

    # Gabungkan semua akun
    akun_list = list(set(saldo_awal_dict.keys()).union(set(df_jurnal["Nama Akun"].unique())))

    # Fungsi bantu untuk pecah saldo ke debit/kredit
    def pecah_saldo_ke_kolom(akun, nilai):
        if akun in akun_debit_normal:
            return (nilai if nilai >= 0 else 0, abs(nilai) if nilai < 0 else 0)
        elif akun in akun_kredit_normal:
            return (abs(nilai) if nilai < 0 else 0, nilai if nilai >= 0 else 0)
        else:
            return (nilai if nilai >= 0 else 0, abs(nilai) if nilai < 0 else 0)

    # Hitung buku besar per akun
    buku_besar = {}
    for akun in akun_list:
        saldo = saldo_awal_dict.get(akun, 0)
        saldo_debit, saldo_kredit = pecah_saldo_ke_kolom(akun, saldo)

        rows = [{
            "Tanggal": "",
            "Keterangan": "Saldo Awal",
            "Debit": "",
            "Kredit": "",
            "Saldo Debit": saldo_debit,
            "Saldo Kredit": saldo_kredit
        }]

        df_akun = df_jurnal[df_jurnal["Nama Akun"] == akun]
        for _, row in df_akun.iterrows():
            saldo = hitung_saldo_berjalan(akun, row["Debit"], row["Kredit"], saldo)
            saldo_debit, saldo_kredit = pecah_saldo_ke_kolom(akun, saldo)

            rows.append({
                "Tanggal": row["Tanggal"].strftime("%Y-%m-%d") if pd.notnull(row["Tanggal"]) else "",
                "Keterangan": row.get("Keterangan", ""),
                "Debit": row["Debit"],
                "Kredit": row["Kredit"],
                "Saldo Debit": saldo_debit,
                "Saldo Kredit": saldo_kredit
            })

        buku_besar[akun] = pd.DataFrame(rows)

    # Format rupiah
    def format_rupiah(x):
        if x == "" or x == 0 or pd.isna(x):
            return ""
        else:
            return f"Rp {int(x):,}".replace(",", ".")

    # Tampilkan hasil
    for akun, df_bb in buku_besar.items():
        st.subheader(f"Nama Akun: {akun}")
        df_display = df_bb.copy()
        df_display["Debit"] = df_display["Debit"].apply(format_rupiah)
        df_display["Kredit"] = df_display["Kredit"].apply(format_rupiah)
        df_display["Saldo Debit"] = df_display["Saldo Debit"].apply(format_rupiah)
        df_display["Saldo Kredit"] = df_display["Saldo Kredit"].apply(format_rupiah)
        st.dataframe(df_display, use_container_width=True)
    
    # Simpan ke sesssion_state agar bisa digunakan di menu lain
    st.session_state["buku_besar"] = buku_besar
    
# Neraca Saldo Setelah Disesuaikan
if menu == "Neraca Saldo Setelah Disesuaikan":
    st.title("🐝 CV Bee Alaska")
    st.subheader("Neraca Saldo Setelah Disesuaikan")
    
    akun_debit_normal = ["Kas", "Piutang Usaha", "Perlengkapan", "Persediaan Pakan", "Persediaan Madu", 
                         "Peralatan usaha", "Kendaraan", "Prive, Alwi", "Harga Pokok Penjualan", 
                         "Beban Gaji Karyawan", "Beban Pakan Lebah", "Beban Perlengkapan", 
                         "Beban Perawatan Sarang", "Beban Penyusutan Kendaraan", "Beban lain-lain", "Beban Pajak"]
    
    akun_kredit_normal = ["Akumulasi Penyusutan Kendaraan", "Utang Usaha", "Utang Gaji", "Utang Pajak", 
                          "Modal, Alwi", "Penjualan"]

    buku_besar = st.session_state.get("buku_besar", {})
    
    rows = []
    total_debit = 0
    total_kredit = 0
    
    for akun, df_bb in buku_besar.items():
        if len(df_bb) > 0:
            saldo_debit = df_bb.iloc[-1]["Saldo Debit"] if not pd.isna(df_bb.iloc[-1]["Saldo Debit"]) else 0
            saldo_kredit = df_bb.iloc[-1]["Saldo Kredit"] if not pd.isna(df_bb.iloc[-1]["Saldo Kredit"]) else 0
            
            if akun in akun_debit_normal:
                debit_val = saldo_debit
                kredit_val = 0
                total_debit += debit_val
            elif akun in akun_kredit_normal:
                debit_val = 0
                kredit_val = saldo_kredit
                total_kredit += kredit_val
            
            if debit_val > 0 or kredit_val > 0:
                rows.append({
                    "Nama Akun": akun,
                    "Debit": debit_val if debit_val > 0 else "",
                    "Kredit": kredit_val if kredit_val > 0 else ""
                })
    
    rows.append({
        "Nama Akun": "TOTAL",
        "Debit": total_debit,
        "Kredit": total_kredit
    })

    # Simpan ke session_state tanpa baris total
    st.session_state["neraca_saldo_disesuaikan"] = rows[:-1]

    def format_rupiah(x):
        if x == "" or x == 0 or pd.isna(x):
            return ""
        else:
            return f"Rp {int(x):,}".replace(",", ".")

    df_display = pd.DataFrame(rows)
    df_display["Debit"] = df_display["Debit"].apply(format_rupiah)
    df_display["Kredit"] = df_display["Kredit"].apply(format_rupiah)
    st.dataframe(df_display, use_container_width=True)
    
    if total_debit != total_kredit:
            st.warning("⚠️ Total Debit tidak sama dengan Total Kredit!")
    
    # Simpan ke Excel

# Laporan Laba Rugi
if menu == "Laporan Laba Rugi":
    st.title("🐝 CV Bee Alaska")
    st.subheader("Laporan Laba Rugi")

    # Ambil data dari session_state
    neraca_saldo = st.session_state.get("neraca_saldo_disesuaikan", [])
    saldo_awal = st.session_state.get("saldo_awal", [])

    def format_rupiah(x):
        if x == 0 or x == "" or pd.isna(x):
            return ""
        return f"Rp {int(x):,}".replace(",", ".")

    if neraca_saldo:
        df_neraca = pd.DataFrame(neraca_saldo)
        df_awal = pd.DataFrame(saldo_awal)

        # Fungsi bantu
        def get_saldo(df, nama, kolom="Debit"):
            row = df[df["Nama Akun"] == nama]
            if not row.empty:
                return row.iloc[0][kolom] if row.iloc[0][kolom] != "" else 0
            return 0

        # Ambil nilai dari neraca disesuaikan dan saldo awal
        penjualan = get_saldo(df_neraca, "Penjualan", "Kredit")
        persediaan_madu_akhir = get_saldo(df_neraca, "Persediaan Madu", "Debit")
        beban_gaji = get_saldo(df_neraca, "Beban Gaji Karyawan")
        beban_pakan = get_saldo(df_neraca, "Beban Pakan Lebah")
        beban_perlengkapan = get_saldo(df_neraca, "Beban Perlengkapan")
        beban_perawatan = get_saldo(df_neraca, "Beban Perawatan Sarang")
        penyusutan = get_saldo(df_neraca, "Beban Penyusutan Kendaraan")
        beban_lain = get_saldo(df_neraca, "Beban lain-lain")
        beban_pajak = get_saldo(df_neraca, "Beban Pajak")
        persediaan_madu_awal = get_saldo(df_awal, "Persediaan Madu")

        # Perhitungan
        biaya_produksi = beban_gaji + beban_pakan + beban_perlengkapan + beban_perawatan + penyusutan
        barang_tersedia = persediaan_madu_awal + biaya_produksi
        hpp = barang_tersedia - persediaan_madu_akhir
        laba_kotor = penjualan - hpp
        beban_operasional_lain = beban_lain
        laba_sebelum_pajak = laba_kotor - beban_operasional_lain
        laba_setelah_pajak = laba_sebelum_pajak - beban_pajak
        
        # Simpan laba bersih ke session_state agar bisa dipakai di laporan perubahan ekuitas
        st.session_state["laba_bersih"] = laba_setelah_pajak

        # Siapkan data untuk ditampilkan dalam 2 kolom
        data = [
            ["Penjualan Bersih", "", "", format_rupiah(penjualan)],
            ["", "", "",""],
            ["Harga Pokok Penjualan (HPP):", "", "", ""],
            ["Persediaan Madu Awal", "", format_rupiah(persediaan_madu_awal), ""],
            ["Harga Pokok Produksi:", "", "", "",],
            ["Beban Pakan Lebah", format_rupiah(beban_pakan), "", ""],
            ["Beban Gaji Karyawan", format_rupiah(beban_gaji), "", ""],
            ["Beban Perlengkapan", format_rupiah(beban_perlengkapan), "", ""],
            ["Beban Perawatan Sarang", format_rupiah(beban_perawatan), "", ""],
            ["Beban Penyusutan Kendaraan", format_rupiah(penyusutan), "", ""],  
            ["Total Harga Pokok Produksi", "", format_rupiah(biaya_produksi), ""],
            ["Barang Tersedia untuk Dijual", "", format_rupiah(barang_tersedia), ""],
            ["Persediaan Madu Akhir",  "", format_rupiah(persediaan_madu_akhir), ""],
            ["Harga Pokok Penjualan",  "", "", format_rupiah(hpp)],
            ["", "", "", "",],
            ["Laba Kotor", "", "", format_rupiah(laba_kotor)],
            ["", "", "", ""],
            ["Beban Operasional:", "", "", ""],
            ["Beban Lain-lain", "", format_rupiah(beban_lain), ""],
            ["Total Beban Operasional", "", "",format_rupiah(beban_operasional_lain)],
            ["", "", "", ""],
            ["Laba Bersih Sebelum Pajak", "", "", format_rupiah(laba_sebelum_pajak)],
            ["Beban Pajak", "", "", format_rupiah(beban_pajak)],
            ["Laba Bersih Setelah Pajak", "", "", format_rupiah(laba_setelah_pajak)],
        ]

        df_laporan = pd.DataFrame(data, columns=["Keterangan", "Jumlah1", "Jumlah2", "Jumlah3"])
        st.dataframe(df_laporan, use_container_width=True, height=700)
    else:
        st.warning("Data Neraca Saldo Disesuaikan belum tersedia. Silakan buka menu tersebut terlebih dahulu.")

    # Simpan ke Excel

# Laporan Perubahan Ekuitas
if menu == "Laporan Perubahan Ekuitas":
    st.title("🐝 CV Bee Alaska")
    st.subheader("Laporan Perubahan Ekuitas")

    # Ambil data dari session_state
    neraca_saldo = st.session_state.get("neraca_saldo_disesuaikan")
    laba_bersih = st.session_state.get("laba_bersih")  # diasumsikan diset dari laporan laba rugi

    def format_rupiah(x):
        if x == 0 or x == "" or pd.isna(x):
            return ""
        return f"Rp {int(x):,}".replace(",", ".")

    # Validasi data
    if neraca_saldo is None:
        st.warning("Data Neraca Saldo Disesuaikan belum tersedia. Silakan buka menu tersebut terlebih dahulu.")
    elif laba_bersih is None:
        st.warning("Laporan Laba Rugi belum tersedia. Silakan buka menu Laporan Laba Rugi terlebih dahulu.")
    else:
        # Cari Modal Awal dari akun "Modal, Alwi"
        modal_awal = 0
        for row in neraca_saldo:
            if row["Nama Akun"] == "Modal, Alwi":
                val = row.get("Kredit", 0)
                if isinstance(val, str):
                    val = int(val.replace("Rp ", "").replace(".", ""))
                modal_awal = val
                break

        modal_akhir = modal_awal + laba_bersih
        
        # Simpan modal akhir ke session_state agar bisa dipakai di laporan posisi keuangan
        st.session_state["modal_akhir"] = modal_akhir
        
        # Tampilkan tabel
        data = [
            ["Modal Awal", format_rupiah(modal_awal)],
            ["Laba/Rugi", format_rupiah(laba_bersih)],
            ["Modal Akhir", format_rupiah(modal_akhir)]
        ]

        df_modal = pd.DataFrame(data, columns=["Keterangan", "Jumlah"])
        st.dataframe(df_modal, use_container_width=True)
    
    # Simpan ke Excel
        
# Laporan Posisi Keuangan 
if menu == "Laporan Posisi Keuangan":
    st.title("🐝 CV Bee Alaska")
    st.subheader("Laporan Posisi Keuangan")

    neraca_saldo = st.session_state.get("neraca_saldo_disesuaikan", [])
    modal_akhir = st.session_state.get("modal_akhir")

    def format_rupiah(x):
        if x == "" or x == 0 or pd.isna(x):
            return ""
        return f"Rp {int(x):,}".replace(",", ".")

    if not neraca_saldo:
        st.warning("Data Neraca Saldo Disesuaikan belum tersedia. Silakan buka menu tersebut terlebih dahulu.")
    elif modal_akhir is None:
        st.warning("Modal Akhir belum tersedia. Silakan buka menu Laporan Perubahan Ekuitas terlebih dahulu.")
    else:
        # Klasifikasi akun
        aset_lancar = ["Kas", "Piutang Usaha", "Perlengkapan", "Persediaan Pakan", "Persediaan Madu"]
        aset_tetap = ["Peralatan usaha", "Kendaraan"]
        akumulasi_penyusutan = ["Akumulasi Penyusutan Kendaraan"]
        liabilitas = ["Utang Usaha", "Utang Gaji", "Utang Pajak"]
        akun_prive = "Prive, Alwi"

        df = pd.DataFrame(neraca_saldo)

        def get_saldo(akun):
            row = df[df["Nama Akun"] == akun]
            if not row.empty:
                debit = row.iloc[0]["Debit"] if row.iloc[0]["Debit"] != "" else 0
                kredit = row.iloc[0]["Kredit"] if row.iloc[0]["Kredit"] != "" else 0
                return int(debit) if int(debit) > 0 else int(kredit)
            return 0

        rows_stafel = []

        # ===== ASET =====
        rows_stafel.append(["ASET", "", ""])
        total_aset_lancar = 0
        rows_stafel.append(["  Aset Lancar", "", ""])
        for akun in aset_lancar:
            saldo = get_saldo(akun)
            total_aset_lancar += saldo
            rows_stafel.append([f"    {akun}", format_rupiah(saldo), ""])
        rows_stafel.append(["  Total Aset Lancar", "", format_rupiah(total_aset_lancar)])

        total_aset_tetap = 0
        rows_stafel.append(["  Aset Tidak Lancar", "", ""])
        for akun in aset_tetap:
            saldo = get_saldo(akun)
            total_aset_tetap += saldo
            rows_stafel.append([f"    {akun}", format_rupiah(saldo), ""])

        for akun in akumulasi_penyusutan:
            saldo = get_saldo(akun)
            total_aset_tetap -= saldo
            rows_stafel.append([f"    {akun}", f"-{format_rupiah(saldo)}", ""])

        rows_stafel.append(["  Total Aset Tidak Lancar", "", format_rupiah(total_aset_tetap)])
        total_aset = total_aset_lancar + total_aset_tetap
        rows_stafel.append(["Total Aset", "", format_rupiah(total_aset)])
        
        rows_stafel.append(["", "", ""]) #Jarak 1 baris kosong
        
        # ===== LIABILITAS =====
        rows_stafel.append(["LIABILITAS DAN EKUITAS", "", ""])
        rows_stafel.append(["  Liabilitas", "", ""])
        total_liabilitas = 0
        for akun in liabilitas:
            saldo = get_saldo(akun)
            total_liabilitas += saldo
            rows_stafel.append([f"    {akun}", format_rupiah(saldo), ""])
        rows_stafel.append(["  Total Liabilitas", "", format_rupiah(total_liabilitas)])

        # ===== EKUITAS =====
        rows_stafel.append(["  Ekuitas", "", ""])
        prive_saldo = get_saldo(akun_prive)
        modal_disesuaikan = modal_akhir - prive_saldo

        rows_stafel.append([f"    Modal, Alwi", format_rupiah(modal_akhir), ""])
        rows_stafel.append([f"    {akun_prive}", f"-{format_rupiah(prive_saldo)}", ""])
        rows_stafel.append(["  Total Ekuitas", "", format_rupiah(modal_disesuaikan)])

        total_passiva = total_liabilitas + modal_disesuaikan
        rows_stafel.append(["Total Liabilitas dan Ekuitas", "", format_rupiah(total_passiva)])

        # Tampilkan sebagai tabel
        df_stafel = pd.DataFrame(rows_stafel, columns=["Keterangan", "1", "2"])
        st.table(df_stafel)

        if total_aset != total_passiva:
            st.warning("⚠️ Total Aset tidak sama dengan Total Liabilitas + Ekuitas!")
            
    # Simpan ke Excel
    if st.button("💾 Simpan Semua Laporan ke Excel"):
        wb = Workbook()
        wb.remove(wb.active)  # Hapus sheet default
    
        # 1. Neraca Saldo Awal
        if "neraca_saldo_awal" in st.session_state:
            ws = wb.create_sheet("Neraca Saldo")
            df = pd.DataFrame(st.session_state["neraca_saldo_awal"])
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

        # 2. Jurnal Umum
        if "jurnal_umum" in st.session_state:
            ws = wb.create_sheet("Jurnal Umum")
            df = pd.DataFrame(st.session_state["jurnal_umum"])
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

        # 3. Neraca Saldo Disesuaikan
        if "neraca_saldo_disesuaikan" in st.session_state:
            ws = wb.create_sheet("Neraca Disesuaikan")
            df = pd.DataFrame(st.session_state["neraca_saldo_disesuaikan"])
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

        # 4. Laporan Laba Rugi
        if "laba_rugi_df" in st.session_state:
            ws = wb.create_sheet("Laporan Laba Rugi")
            df = st.session_state["laba_rugi_df"]
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

        # 5. Laporan Perubahan Ekuitas
        if "modal_akhir" in st.session_state and "laba_bersih" in st.session_state:
            ws = wb.create_sheet("Perubahan Ekuitas")
            modal_awal = 0
            for row in st.session_state["neraca_saldo_disesuaikan"]:
                if row["Nama Akun"] == "Modal, Alwi":
                    val = row.get("Kredit", 0)
                    if isinstance(val, str):
                        val = int(val.replace("Rp ", "").replace(".", ""))
                    modal_awal = val
                    break
            laba_bersih = st.session_state["laba_bersih"]
            modal_akhir = st.session_state["modal_akhir"]
            df = pd.DataFrame([
                ["Modal Awal", modal_awal],
                ["Laba/Rugi", laba_bersih],
                ["Modal Akhir", modal_akhir],
            ], columns=["Keterangan", "Jumlah"])
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

        # 6. Laporan Posisi Keuangan
        if "modal_akhir" in st.session_state:
            ws = wb.create_sheet("Posisi Keuangan")
            rows = st.session_state.get("laporan_posisi_keuangan", [])
            df = pd.DataFrame(rows, columns=["Keterangan", "1", "2"])
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

        # Simpan ke bytes dan buat tombol unduh
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        st.download_button(
            label="📥 Download Excel",
            data=buffer,
            file_name="Laporan_Keuangan_CV_Bee_Alaska.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )